"""
The report: a view over the database, rebuildable at any time.

Nothing here queries DNS. Everything is streamed out of SQLite one batch
at a time and collapsed into runs on the way, so a database with tens of
millions of rows reports in roughly constant memory.

Three shapes of output, any combination:

    xlsx  Overview / Prefixes / IP Details / Ranges (+ IPv6 sheets)
    csv   the same sheets as separate files, no row ceiling
    json  a summary document plus newline-delimited detail records
"""

import csv
import json
import time

import openpyxl

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .console import console, good, info
from .store import counter, get_json, get_state, queue_size
from .util import count_cell, fmt, key_to_int, summarize_run


# Hard ceiling of the .xlsx format, header row included.
EXCEL_MAX_ROWS = 1_048_575

# Rows pulled from SQLite at a time.
FETCH_SIZE = 10_000


# =====================
# STREAMING THE RESULTS
# =====================

DETAIL_SQL = """
    SELECT ip_key, version, ip, prefix, hostname, label, verified
    FROM results
    WHERE hostname <> '' {extra}
    ORDER BY label, version, prefix, ip_key
"""


def iter_runs(conn, only_matching=False):
    """
    Stream results as maximal runs of consecutive addresses that share a
    (label, prefix). Only one run is ever held in memory.
    """

    cur = conn.cursor()
    cur.arraysize = FETCH_SIZE
    cur.execute(DETAIL_SQL.format(extra="AND matched = 1" if only_matching
                                  else ""))

    run = []
    key = None
    last_int = None

    while True:

        batch = cur.fetchmany(FETCH_SIZE)

        if not batch:
            break

        for row in batch:

            value = key_to_int(row[0])

            k = (row[5], row[3], row[1])            # label, prefix, version

            if run and (k != key or value != last_int + 1):
                yield key, run
                run = []

            key = k
            last_int = value

            run.append((value,) + tuple(row[1:]))

    if run:
        yield key, run


def iter_report(conn, only_matching=False, verified=False):
    """Yield ('detail'|'range', row) for the whole report."""

    for (label, prefix, version), run in iter_runs(conn, only_matching):

        blocks = summarize_run(run[0][0], run[-1][0], version)

        for block in blocks:
            yield "range", [
                label,
                prefix,
                str(block),
                str(block.network_address),
                str(block.broadcast_address),
                count_cell(block.num_addresses),
            ]

        # blocks are sorted and disjoint, so walk both lists once
        idx = 0

        for value, ver, ip, pfx, hostname, _label, verify_flag in run:

            while (idx < len(blocks)
                   and value > int(blocks[idx].broadcast_address)):
                idx += 1

            block = blocks[idx] if idx < len(blocks) else None

            line = [
                label,
                ip,
                f"IPv{ver}",
                pfx,
                hostname,
                str(block) if block else "",
                count_cell(block.num_addresses) if block else 0,
            ]

            if verified:
                line.append({1: "yes", 0: "NO"}.get(verify_flag, "n/a"))

            yield "detail", line


# =====================
# FIGURES
# =====================

def prefix_rows(conn):
    """One row per announced prefix, with what the scan found in it."""

    seen = {
        r[0]: (r[1], r[2])
        for r in conn.execute(
            "SELECT prefix, COUNT(*), SUM(status = 'ok') "
            "FROM results GROUP BY prefix")
    }

    done = {
        r[0]: r[1]
        for r in conn.execute("SELECT prefix, done FROM progress")
    }

    out = []

    for (prefix, version, first, last, addresses, first_seen, last_seen,
         state) in conn.execute(
            "SELECT prefix, version, first_ip, last_ip, addresses, "
            "first_seen, last_seen, state FROM prefixes "
            "ORDER BY version, prefix"):

        scanned, found = seen.get(prefix, (0, 0))

        out.append([
            prefix,
            f"IPv{version}",
            state,
            first_seen[:10],
            last_seen[:10],
            first,
            last,
            count_cell(int(addresses)),
            scanned,
            found or 0,
            {1: "complete", 0: "partial"}.get(done.get(prefix), "not started"),
        ])

    return out


def prefix_state_counts(conn):

    return dict(conn.execute(
        "SELECT state, COUNT(*) FROM prefixes GROUP BY state"))


V6_NETWORK_HEADER = ["Network In Use", "Announced Prefix", "Nibbles"]

V6_WILDCARD_HEADER = ["Covers", "Announced Prefix", "Wildcard Hostname",
                      "Nibbles"]


def v6_network_rows(conn):
    """Networks the walk proved are in use but was told not to descend."""

    return conn.execute(
        "SELECT network, prefix, depth FROM v6_subnets ORDER BY network"
    ).fetchall()


def v6_wildcard_rows(conn):

    return conn.execute(
        "SELECT covers, prefix, hostname, depth FROM v6_wildcards "
        "ORDER BY covers").fetchall()


def label_rows(conn, only_matching=False, top=None):

    sql = ("SELECT label, COUNT(*) FROM results WHERE hostname <> '' "
           + ("AND matched = 1 " if only_matching else "")
           + "GROUP BY label ORDER BY COUNT(*) DESC")

    rows = conn.execute(sql).fetchall()

    return rows[:top] if top else rows


def overview_rows(conn, args):
    """The Overview sheet: what was scanned, and what came back."""

    asn = get_state(conn, "asn", "?")

    meta = get_json(conn, "asn_info", {}) or {}

    # scoped to IPv4: the sweep's figures must not silently absorb the
    # addresses the IPv6 walk contributed
    counts = dict(conn.execute(
        "SELECT status, COUNT(*) FROM results WHERE version = 4 "
        "GROUP BY status"))

    v4_scope = 0
    v4_prefixes = 0
    v6_prefixes = 0
    v6_scope = 0

    for prefix, version, addresses in conn.execute(
            "SELECT prefix, version, addresses FROM prefixes"):

        if version == 4:
            v4_prefixes += 1
            v4_scope += int(addresses)
        else:
            v6_prefixes += 1
            v6_scope += int(addresses)

    v4_scanned = conn.execute(
        "SELECT COUNT(*) FROM results WHERE version = 4").fetchone()[0]

    states = prefix_state_counts(conn)

    window = get_json(conn, "prefix_window", {}) or {}

    # results sitting in space the ASN no longer announces: still real
    # findings, but no longer part of this network
    stale_hits = conn.execute(
        "SELECT COUNT(*) FROM results r JOIN prefixes p ON r.prefix = p.prefix "
        "WHERE p.state <> 'live' AND r.hostname <> ''").fetchone()[0]

    named = conn.execute(
        "SELECT COUNT(*) FROM results WHERE hostname <> ''").fetchone()[0]

    v6_found = conn.execute(
        "SELECT COUNT(*) FROM results WHERE version = 6").fetchone()[0]

    matched = conn.execute(
        "SELECT COUNT(*) FROM results WHERE matched = 1 AND hostname <> ''"
    ).fetchone()[0]

    distinct_labels = conn.execute(
        "SELECT COUNT(DISTINCT label) FROM results WHERE hostname <> ''"
    ).fetchone()[0]

    v6_nodes = counter(conn, "v6_nodes")
    v6_subnets = conn.execute("SELECT COUNT(*) FROM v6_subnets").fetchone()[0]
    v6_wild = conn.execute("SELECT COUNT(*) FROM v6_wildcards").fetchone()[0]
    v6_dead = conn.execute("SELECT COUNT(*) FROM v6_dead").fetchone()[0]
    frontier = queue_size(conn)

    rows = [
        # first line, so a report can never be mistaken for another network
        ("ASN", f"AS{asn}"),
        ("Holder", meta.get("holder") or "unknown"),
        ("Country", meta.get("country") or ""),
        ("Registry", meta.get("registry") or ""),
        ("Prefix source", meta.get("prefix_source") or meta.get("source", "")),
        ("Report generated", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),

        ("ANNOUNCED SPACE", ""),
        ("Announcement window",
         f"{window['start'][:10]} to {window['end'][:10]}"
         if window.get("start") and window.get("end") else "current table"),
        ("IPv4 prefixes", v4_prefixes),
        ("IPv4 addresses", count_cell(v4_scope)),
        ("IPv6 prefixes", v6_prefixes),
        ("IPv6 addresses", count_cell(v6_scope)),
        ("  still announced (live)", states.get("live", 0)),
        ("  seen in window, gone by its end (recent)", states.get("recent", 0)),
        ("  no longer announced (withdrawn)", states.get("withdrawn", 0)),
        ("", ""),

        ("IPv4 SWEEP", ""),
        ("Addresses probed", v4_scanned),
        ("Coverage",
         f"{v4_scanned / v4_scope * 100:.2f}%" if v4_scope else "n/a"),
        ("PTR found", counts.get("ok", 0)),
        ("No PTR", counts.get("nxdomain", 0)),
        ("Unresolved (timeout / SERVFAIL)", counts.get("error", 0)),
        ("Skipped, no reverse zone", counts.get("no_zone", 0)),
        ("", ""),

        ("IPv6 TREE WALK", ""),
        ("ip6.arpa nodes queried", v6_nodes),
        ("Addresses found", v6_found),
        ("Networks in use (stopped at max depth)", v6_subnets),
        ("Wildcard reverse zones", v6_wild),
        ("Nodes that never answered", v6_dead),
        ("Frontier still queued", frontier),
        ("", ""),

        ("RESULTS", ""),
        ("Addresses with a hostname", named),
        ("Matched by rules", matched if args.rules else "n/a (no rules file)"),
        ("Distinct labels", distinct_labels),
        ("", ""),

        ("BREAKDOWN", ""),
    ]

    for label, n in label_rows(conn, args.only_matching, top=40):
        rows.append((f"  {label}", n))

    unresolved = counts.get("error", 0) + counts.get("no_zone", 0)

    if unresolved:
        rows += [
            ("", ""),
            ("NOTE", f"{fmt(unresolved)} IPv4 addresses were never answered "
                     f"and may hide hosts. Re-run with --retry-errors "
                     f"[--include-dead]."),
        ]

    if frontier:
        rows += [
            ("", ""),
            ("NOTE", f"{fmt(frontier)} ip6.arpa nodes are still queued. "
                     f"Re-run with a larger --v6-max-nodes to finish."),
        ]

    if stale_hits:
        rows += [
            ("", ""),
            ("NOTE", f"{fmt(stale_hits)} of the addresses below sit in "
                     f"blocks the ASN no longer announces (marked 'recent' "
                     f"or 'withdrawn' on the Prefixes sheet). They were "
                     f"found when that space was still routed here."),
        ]

    return rows


# =====================
# EXCEL
# =====================

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FONT = Font(bold=True)

DETAIL_HEADER = ["Label", "IP", "Family", "Announced Prefix", "Hostname",
                 "Contiguous Range", "Range Size"]
DETAIL_WIDTHS = [26, 40, 8, 24, 56, 30, 14, 16]

RANGE_HEADER = ["Label", "Announced Prefix", "Contiguous Range",
                "First IP", "Last IP", "Addresses"]
RANGE_WIDTHS = [26, 24, 30, 40, 40, 14]

PREFIX_HEADER = ["Prefix", "Family", "Announcement", "First Seen",
                 "Last Seen", "First IP", "Last IP", "Addresses", "Probed",
                 "With PTR", "Scan Status"]
PREFIX_WIDTHS = [24, 8, 14, 12, 12, 40, 40, 22, 14, 12, 14]


class SheetWriter:
    """
    Write-only sheet that spills into "<title> (2)", "(3)"... when the
    xlsx row ceiling is reached, re-emitting the header each time.
    """

    def __init__(self, wb, title, header, widths, max_rows):

        self.wb = wb
        self.title = title
        self.header = header
        self.widths = widths
        self.max_rows = max_rows

        self.part = 0
        self.rows = 0
        self.total = 0
        self.ws = None

        self._new_sheet()

    def _new_sheet(self):

        if self.ws is not None:
            self._finish_sheet()

        self.part += 1

        name = self.title if self.part == 1 else f"{self.title} ({self.part})"

        self.ws = self.wb.create_sheet(name)

        # write_only emits sheetViews before the rows, so the frozen
        # header has to be set now, not at close()
        self.ws.freeze_panes = "A2"

        for i, width in enumerate(self.widths[:len(self.header)], start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = width

        cells = []

        for value in self.header:
            cell = WriteOnlyCell(self.ws, value=value)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center")
            cells.append(cell)

        self.ws.append(cells)

        self.rows = 0

    def _finish_sheet(self):

        # the filter range is only known once the sheet is full, and
        # autoFilter is serialised after the rows, so it belongs here
        self.ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(self.header))}{self.rows + 1}")

    def append(self, row):

        if self.rows >= self.max_rows:
            self._new_sheet()

        self.ws.append(row)

        self.rows += 1
        self.total += 1

    def close(self):
        self._finish_sheet()


def save_excel(conn, path, args):

    verified = args.verify

    detail_header = list(DETAIL_HEADER)

    if verified:
        detail_header.append("Forward Confirmed")

    # write_only streams rows to disk instead of building the whole
    # workbook in memory
    wb = openpyxl.Workbook(write_only=True)

    overview = wb.create_sheet("Overview")
    overview.column_dimensions["A"].width = 42
    overview.column_dimensions["B"].width = 70

    cells = []

    for value in ("Metric", "Value"):
        cell = WriteOnlyCell(overview, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cells.append(cell)

    overview.append(cells)

    for name, value in overview_rows(conn, args):

        if value == "" and name and not name.startswith(" "):
            cell = WriteOnlyCell(overview, value=name)
            cell.font = SECTION_FONT
            overview.append([cell, ""])
        else:
            overview.append([name, value])

    prefixes = SheetWriter(wb, "Prefixes", PREFIX_HEADER, PREFIX_WIDTHS,
                           args.max_rows)

    for row in prefix_rows(conn):
        prefixes.append(row)

    prefixes.close()

    detail = SheetWriter(wb, "IP Details", detail_header, DETAIL_WIDTHS,
                         args.max_rows)

    ranges = SheetWriter(wb, "Ranges", RANGE_HEADER, RANGE_WIDTHS,
                         args.max_rows)

    for kind, row in iter_report(conn, args.only_matching, verified):

        if kind == "detail":
            detail.append(row)
        else:
            ranges.append(row)

    detail.close()
    ranges.close()

    v6_extra(wb, conn, args)

    # the writers create spill sheets as they fill, which interleaves
    # them; group each family back together before serialising
    order = {"Overview": 0, "Prefixes": 1, "IP Details": 2, "Ranges": 3,
             "IPv6 Networks": 4, "IPv6 Wildcards": 5}

    wb._sheets.sort(key=lambda s: (order.get(s.title.split(" (")[0], 9),
                                   s.title))

    wb.save(path)

    good(f"{path}  ({fmt(detail.total)} addresses, {fmt(ranges.total)} "
         f"ranges" + (f", spilled over {detail.part} sheets"
                      if detail.part > 1 else "") + ")")

    return detail.total


def v6_extra(wb, conn, args):
    """Sheets that only exist when the IPv6 walk produced something."""

    for title, header, widths, rows in (
            ("IPv6 Networks", V6_NETWORK_HEADER, [44, 24, 10],
             v6_network_rows(conn)),
            ("IPv6 Wildcards", V6_WILDCARD_HEADER, [44, 24, 56, 10],
             v6_wildcard_rows(conn))):

        if not rows:
            continue

        ws = SheetWriter(wb, title, header, widths, args.max_rows)

        for row in rows:
            ws.append(list(row))

        ws.close()


# =====================
# CSV
# =====================

def save_csv(conn, path, args):

    base = path[:-5] if path.lower().endswith(".xlsx") else path

    detail_header = list(DETAIL_HEADER)

    if args.verify:
        detail_header.append("Forward Confirmed")

    n_detail = 0
    n_range = 0

    # utf-8-sig so Excel opens the file with the right encoding
    with open(f"{base}_addresses.csv", "w", newline="",
              encoding="utf-8-sig") as fd, \
         open(f"{base}_ranges.csv", "w", newline="",
              encoding="utf-8-sig") as fr:

        wd = csv.writer(fd)
        wr = csv.writer(fr)

        wd.writerow(detail_header)
        wr.writerow(RANGE_HEADER)

        for kind, row in iter_report(conn, args.only_matching, args.verify):

            if kind == "detail":
                wd.writerow(row)
                n_detail += 1
            else:
                wr.writerow(row)
                n_range += 1

    with open(f"{base}_prefixes.csv", "w", newline="",
              encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(PREFIX_HEADER)
        w.writerows(prefix_rows(conn))

    with open(f"{base}_overview.csv", "w", newline="",
              encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Metric", "Value"])
        w.writerows(overview_rows(conn, args))

    good(f"{base}_addresses.csv  ({fmt(n_detail)} addresses)")
    good(f"{base}_ranges.csv  ({fmt(n_range)} ranges)")
    good(f"{base}_prefixes.csv")
    good(f"{base}_overview.csv")

    # written only when the IPv6 walk actually produced them, so a run
    # that found none does not leave a trail of empty files
    for suffix, header, rows in (
            ("ipv6_networks", V6_NETWORK_HEADER, v6_network_rows(conn)),
            ("ipv6_wildcards", V6_WILDCARD_HEADER, v6_wildcard_rows(conn))):

        if not rows:
            continue

        with open(f"{base}_{suffix}.csv", "w", newline="",
                  encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)

        good(f"{base}_{suffix}.csv  ({fmt(len(rows))} rows)")

    return n_detail


# =====================
# JSON
# =====================

def save_json(conn, path, args):
    """A summary document, plus one JSON object per address on its own line."""

    base = path[:-5] if path.lower().endswith(".xlsx") else path

    summary_path = f"{base}.json"
    detail_path = f"{base}_addresses.jsonl"

    ranges = []
    n_detail = 0

    detail_header = list(DETAIL_HEADER)

    if args.verify:
        detail_header.append("Forward Confirmed")

    keys = ["label", "ip", "family", "prefix", "hostname", "range",
            "range_size", "forward_confirmed"]

    with open(detail_path, "w", encoding="utf-8") as f:

        for kind, row in iter_report(conn, args.only_matching, args.verify):

            if kind == "detail":
                f.write(json.dumps(dict(zip(keys, row))) + "\n")
                n_detail += 1

            else:
                ranges.append({
                    "label": row[0], "prefix": row[1], "range": row[2],
                    "first_ip": row[3], "last_ip": row[4],
                    "addresses": row[5],
                })

    summary = {
        "asn": int(get_state(conn, "asn", 0)),
        "generated": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "asn_info": get_json(conn, "asn_info", {}),
        "overview": {k: v for k, v in overview_rows(conn, args) if k and v != ""},
        "prefixes": [
            dict(zip(["prefix", "family", "announcement", "first_seen",
                      "last_seen", "first_ip", "last_ip", "addresses",
                      "probed", "with_ptr", "scan_status"], row))
            for row in prefix_rows(conn)
        ],
        "labels": [{"label": label, "addresses": n}
                   for label, n in label_rows(conn, args.only_matching)],
        "ipv6_networks_in_use": [
            {"network": r[0], "prefix": r[1], "nibbles": r[2]}
            for r in v6_network_rows(conn)
        ],
        "ipv6_wildcards": [
            {"covers": r[0], "prefix": r[1], "hostname": r[2], "nibbles": r[3]}
            for r in v6_wildcard_rows(conn)
        ],
        "ranges": ranges,
        "addresses_file": detail_path,
        "address_count": n_detail,
    }

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    good(f"{summary_path}  ({len(ranges)} ranges)")
    good(f"{detail_path}  ({fmt(n_detail)} addresses)")

    return n_detail


# =====================
# CONSOLE SUMMARY
# =====================

def print_summary(conn, args):
    """The at-a-glance table printed when a run finishes."""

    from rich.table import Table

    labels = label_rows(conn, args.only_matching, top=15)

    if not labels:
        info("No addresses with a PTR record were found yet.")
        return

    table = Table(title="Top labels", title_style="bold",
                  header_style="head", box=None, pad_edge=False)

    table.add_column("Label")
    table.add_column("Addresses", justify="right")

    for label, n in labels:
        table.add_row(label, fmt(n))

    console.print()
    console.print(table)
    console.print()


def save_report(conn, args):

    formats = (["xlsx", "csv", "json"] if args.format == "all"
               else [args.format])

    for fmt_name in formats:

        if fmt_name == "xlsx":
            save_excel(conn, args.output, args)

        elif fmt_name == "csv":
            save_csv(conn, args.output, args)

        elif fmt_name == "json":
            save_json(conn, args.output, args)
